from flask import Blueprint, render_template, redirect, url_for, flash, request, Response
from flask_login import login_required, current_user
from ..models import Ticket, TicketExpense, SparePart, StockTransaction
from ..utils import to_buddhist_era
from .. import db
from sqlalchemy import func, extract
from functools import wraps
from datetime import datetime
import io

finance_bp = Blueprint('finance', __name__, url_prefix='/finance')


def admin_manager_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role not in ['Admin', 'Manager']:
            flash('คุณไม่มีสิทธิ์เข้าถึงหน้านี้', 'danger')
            return redirect(url_for('dashboard.index'))
        return f(*args, **kwargs)
    return decorated_function


# ==========================================
# ค่าใช้จ่ายต่อ Ticket
# ==========================================

@finance_bp.route('/ticket/<int:ticket_id>/expense', methods=['POST'])
@login_required
def add_expense(ticket_id):
    """เพิ่มค่าใช้จ่ายให้ Ticket"""
    ticket = Ticket.query.get_or_404(ticket_id)

    expense_type = request.form.get('expense_type', 'other')
    description = request.form.get('description', '').strip()
    amount = float(request.form.get('amount', 0) or 0)
    quantity = int(request.form.get('quantity', 1) or 1)
    receipt_number = request.form.get('receipt_number', '').strip()

    if not description or amount <= 0:
        flash('กรุณากรอกรายละเอียดและจำนวนเงินให้ถูกต้อง', 'danger')
        return redirect(url_for('tickets.ticket_detail', ticket_id=ticket_id))

    expense = TicketExpense(
        ticket_id=ticket_id,
        expense_type=expense_type,
        description=description,
        amount=amount,
        quantity=quantity,
        total=amount * quantity,
        receipt_number=receipt_number,
        created_by=current_user.id,
    )
    db.session.add(expense)
    db.session.commit()

    flash(f'เพิ่มค่าใช้จ่าย ฿{amount * quantity:,.2f} เรียบร้อย', 'success')
    return redirect(url_for('tickets.ticket_detail', ticket_id=ticket_id))


@finance_bp.route('/expense/<int:expense_id>/delete', methods=['POST'])
@login_required
def delete_expense(expense_id):
    """ลบค่าใช้จ่าย"""
    expense = TicketExpense.query.get_or_404(expense_id)
    ticket_id = expense.ticket_id

    if expense.created_by != current_user.id and current_user.role != 'Admin':
        flash('คุณไม่มีสิทธิ์ลบรายการนี้', 'danger')
        return redirect(url_for('tickets.ticket_detail', ticket_id=ticket_id))

    db.session.delete(expense)
    db.session.commit()
    flash('ลบรายการค่าใช้จ่ายเรียบร้อย', 'success')
    return redirect(url_for('tickets.ticket_detail', ticket_id=ticket_id))


# ==========================================
# รายงานค่าใช้จ่าย
# ==========================================

@finance_bp.route('/report')
@login_required
@admin_manager_required
def expense_report():
    """หน้ารายงานค่าใช้จ่าย — สรุปรายเดือน/ปี"""
    year = request.args.get('year', datetime.utcnow().year, type=int)
    month = request.args.get('month', 0, type=int)  # 0 = ทุกเดือน

    # ค่าใช้จ่ายรายเดือนตลอดทั้งปี
    monthly_data = db.session.query(
        extract('month', TicketExpense.created_at).label('month'),
        func.sum(TicketExpense.total).label('total')
    ).filter(
        extract('year', TicketExpense.created_at) == year
    ).group_by(
        extract('month', TicketExpense.created_at)
    ).all()

    monthly_totals = {int(m): float(t or 0) for m, t in monthly_data}

    # ค่าใช้จ่ายแยกตามประเภท
    type_query = db.session.query(
        TicketExpense.expense_type,
        func.sum(TicketExpense.total).label('total'),
        func.count(TicketExpense.id).label('count')
    ).filter(extract('year', TicketExpense.created_at) == year)

    if month > 0:
        type_query = type_query.filter(extract('month', TicketExpense.created_at) == month)

    type_data = type_query.group_by(TicketExpense.expense_type).all()

    # Top 10 Ticket ที่มีค่าใช้จ่ายสูงสุด
    top_query = db.session.query(
        Ticket.id,
        Ticket.ticket_number,
        Ticket.title,
        func.sum(TicketExpense.total).label('total_cost')
    ).join(TicketExpense).filter(
        extract('year', TicketExpense.created_at) == year
    )
    if month > 0:
        top_query = top_query.filter(extract('month', TicketExpense.created_at) == month)

    top_tickets = top_query.group_by(Ticket.id).order_by(func.sum(TicketExpense.total).desc()).limit(10).all()

    # ค่าใช้จ่ายรวมทั้งปี / เดือนที่เลือก
    total_year = sum(monthly_totals.values())

    # รายการค่าใช้จ่ายล่าสุด
    recent_query = TicketExpense.query.filter(
        extract('year', TicketExpense.created_at) == year
    )
    if month > 0:
        recent_query = recent_query.filter(extract('month', TicketExpense.created_at) == month)
    recent_expenses = recent_query.order_by(TicketExpense.created_at.desc()).limit(20).all()

    # รายปีที่มีข้อมูล
    available_years = db.session.query(
        extract('year', TicketExpense.created_at).label('year')
    ).distinct().order_by(extract('year', TicketExpense.created_at).desc()).all()
    available_years = [int(y[0]) for y in available_years] if available_years else [datetime.utcnow().year]

    thai_months = ['', 'ม.ค.', 'ก.พ.', 'มี.ค.', 'เม.ย.', 'พ.ค.', 'มิ.ย.',
                   'ก.ค.', 'ส.ค.', 'ก.ย.', 'ต.ค.', 'พ.ย.', 'ธ.ค.']

    return render_template('expense_report.html',
                           year=year,
                           month=month,
                           monthly_totals=monthly_totals,
                           type_data=type_data,
                           top_tickets=top_tickets,
                           total_year=total_year,
                           recent_expenses=recent_expenses,
                           available_years=available_years,
                           thai_months=thai_months,
                           expense_types=TicketExpense.EXPENSE_TYPES)


@finance_bp.route('/report/print')
@login_required
@admin_manager_required
def print_report():
    """หน้าพิมพ์รายงานค่าใช้จ่ายสำหรับผู้บริหาร"""
    year = request.args.get('year', datetime.utcnow().year, type=int)
    month = request.args.get('month', 0, type=int)

    # ดึงข้อมูลเหมือน expense_report
    monthly_data = db.session.query(
        extract('month', TicketExpense.created_at).label('month'),
        func.sum(TicketExpense.total).label('total')
    ).filter(
        extract('year', TicketExpense.created_at) == year
    ).group_by(extract('month', TicketExpense.created_at)).all()
    monthly_totals = {int(m): float(t or 0) for m, t in monthly_data}

    type_query = db.session.query(
        TicketExpense.expense_type,
        func.sum(TicketExpense.total).label('total'),
        func.count(TicketExpense.id).label('count')
    ).filter(extract('year', TicketExpense.created_at) == year)
    if month > 0:
        type_query = type_query.filter(extract('month', TicketExpense.created_at) == month)
    type_data = type_query.group_by(TicketExpense.expense_type).all()

    top_query = db.session.query(
        Ticket.id, Ticket.ticket_number, Ticket.title,
        func.sum(TicketExpense.total).label('total_cost')
    ).join(TicketExpense).filter(extract('year', TicketExpense.created_at) == year)
    if month > 0:
        top_query = top_query.filter(extract('month', TicketExpense.created_at) == month)
    top_tickets = top_query.group_by(Ticket.id).order_by(func.sum(TicketExpense.total).desc()).limit(20).all()

    all_query = TicketExpense.query.filter(extract('year', TicketExpense.created_at) == year)
    if month > 0:
        all_query = all_query.filter(extract('month', TicketExpense.created_at) == month)
    all_expenses = all_query.order_by(TicketExpense.created_at.desc()).all()

    total_year = sum(monthly_totals.values())
    thai_months = ['', 'มกราคม', 'กุมภาพันธ์', 'มีนาคม', 'เมษายน', 'พฤษภาคม', 'มิถุนายน',
                   'กรกฎาคม', 'สิงหาคม', 'กันยายน', 'ตุลาคม', 'พฤศจิกายน', 'ธันวาคม']

    return render_template('expense_print.html',
                           year=year, month=month,
                           monthly_totals=monthly_totals,
                           type_data=type_data,
                           top_tickets=top_tickets,
                           all_expenses=all_expenses,
                           total_year=total_year,
                           thai_months=thai_months,
                           expense_types=TicketExpense.EXPENSE_TYPES,
                           now=datetime.utcnow(),
                           to_buddhist_era=to_buddhist_era)


@finance_bp.route('/report/export')
@login_required
@admin_manager_required
def export_expenses():
    """Export ค่าใช้จ่ายเป็น Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, numbers

    year = request.args.get('year', datetime.utcnow().year, type=int)
    month = request.args.get('month', 0, type=int)

    query = TicketExpense.query.filter(extract('year', TicketExpense.created_at) == year)
    if month > 0:
        query = query.filter(extract('month', TicketExpense.created_at) == month)
    expenses = query.order_by(TicketExpense.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"

    headers = ['วันที่', 'เลข Ticket', 'หัวข้อ Ticket', 'ประเภท', 'รายละเอียด',
               'ราคาต่อหน่วย', 'จำนวน', 'รวม', 'เลขใบเสร็จ', 'ผู้บันทึก']
    header_fill = PatternFill(start_color='B91C1C', end_color='B91C1C', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row, e in enumerate(expenses, 2):
        ws.cell(row=row, column=1, value=to_buddhist_era(e.created_at))
        ws.cell(row=row, column=2, value=e.ticket.ticket_number if e.ticket else '-')
        ws.cell(row=row, column=3, value=e.ticket.title if e.ticket else '-')
        ws.cell(row=row, column=4, value=e.get_type_display())
        ws.cell(row=row, column=5, value=e.description)
        ws.cell(row=row, column=6, value=e.amount)
        ws.cell(row=row, column=7, value=e.quantity)
        ws.cell(row=row, column=8, value=e.total)
        ws.cell(row=row, column=9, value=e.receipt_number or '-')
        ws.cell(row=row, column=10, value=e.creator.name if e.creator else '-')

    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_length + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=expenses_{year}_{month or "all"}.xlsx'}
    )
