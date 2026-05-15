from flask import Blueprint, render_template, redirect, url_for, flash, request, Response, current_app, send_from_directory
from flask_login import login_required, current_user
from ..models import Ticket, TicketLog, TicketComment, Category, User, SystemConfig, FileAttachment, SparePart, TicketExpense, Equipment
from ..utils import get_allowed_statuses, send_notification, generate_qr_base64
from .. import db
from datetime import datetime
import csv
import io
import os
import uuid

tickets_bp = Blueprint('tickets', __name__, url_prefix='/tickets')

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'pdf', 'doc', 'docx', 'xls', 'xlsx', 'txt', 'zip', 'rar'}


def get_branch_options():
    raw = SystemConfig.get('branch_options', '')
    options = [line.strip() for line in raw.splitlines() if line.strip()]
    return options


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@tickets_bp.route('/')
@login_required
def ticket_list():
    """หน้ารายการ Ticket ทั้งหมด + Search/Filter"""
    page = request.args.get('page', 1, type=int)
    per_page = 15

    search = request.args.get('search', '').strip()
    status_filter = request.args.get('status', '')
    priority_filter = request.args.get('priority', '')
    category_filter = request.args.get('category', '', type=int) if request.args.get('category') else None

    query = Ticket.query

    if search:
        query = query.filter(
            db.or_(
                Ticket.title.ilike(f'%{search}%'),
                Ticket.ticket_number.ilike(f'%{search}%'),
                Ticket.description.ilike(f'%{search}%'),
            )
        )
    if status_filter:
        query = query.filter_by(status=status_filter)
    if priority_filter:
        query = query.filter_by(priority=priority_filter)
    if category_filter:
        query = query.filter_by(category_id=category_filter)

    tickets = query.order_by(Ticket.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    categories = Category.query.all()

    return render_template('ticket_list.html',
                           tickets=tickets,
                           categories=categories,
                           search=search,
                           status_filter=status_filter,
                           priority_filter=priority_filter,
                           category_filter=category_filter)


@tickets_bp.route('/new', methods=['GET', 'POST'])
@login_required
def ticket_create():
    """สร้าง Ticket ใหม่"""
    branch_options = get_branch_options()
    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        description = request.form.get('description', '').strip()
        contact_phone = request.form.get('contact_phone', '').strip()
        branch_location = request.form.get('branch_location', '').strip()
        priority = request.form.get('priority', 'Medium')
        category_id = request.form.get('category_id', type=int)
        equipment_id = request.form.get('equipment_id', type=int)

        if not title:
            flash('กรุณากรอกหัวข้อปัญหา', 'danger')
            return redirect(url_for('tickets.ticket_create'))
        if not contact_phone:
            flash('กรุณากรอกเบอร์ติดต่อ', 'danger')
            return redirect(url_for('tickets.ticket_create'))
        if not branch_location or branch_location not in branch_options:
            flash('กรุณาเลือกสาขาจากรายการที่กำหนด', 'danger')
            return redirect(url_for('tickets.ticket_create'))

        normalized_phone = ''.join(ch for ch in contact_phone if ch.isdigit())
        if len(normalized_phone) < 9:
            flash('เบอร์ติดต่อไม่ถูกต้อง', 'danger')
            return redirect(url_for('tickets.ticket_create'))

        full_description = (
            f"สาขา: {branch_location}\n"
            f"เบอร์ติดต่อ: {contact_phone}\n\n"
            f"{description}"
        ).strip()

        ticket = Ticket(
            ticket_number=Ticket.generate_ticket_number(),
            title=title,
            description=full_description,
            priority=priority,
            category_id=category_id,
            equipment_id=equipment_id,
            requester_id=current_user.id,
            status='Pending',
        )
        db.session.add(ticket)
        db.session.flush()

        log = TicketLog(
            ticket_id=ticket.id,
            actor_id=current_user.id,
            action='สร้างใบแจ้งซ่อม',
            new_value='Pending',
        )
        db.session.add(log)

        # จัดการ File Upload
        files = request.files.getlist('attachments')
        for f in files:
            if f and f.filename and allowed_file(f.filename):
                ext = f.filename.rsplit('.', 1)[1].lower()
                safe_name = f"{uuid.uuid4().hex}.{ext}"
                save_path = os.path.join(current_app.config['UPLOAD_FOLDER'], safe_name)
                f.save(save_path)

                attachment = FileAttachment(
                    ticket_id=ticket.id,
                    filename=safe_name,
                    original_name=f.filename,
                    file_size=os.path.getsize(save_path),
                    mime_type=f.content_type,
                    uploaded_by=current_user.id,
                )
                db.session.add(attachment)

        db.session.commit()

        # ส่งการแจ้งเตือน
        try:
            config = SystemConfig.get_notification_config()
            send_notification(ticket, f'สร้างใบแจ้งซ่อมใหม่ โดย {current_user.name}', config)
        except Exception:
            pass

        flash(f'สร้างใบแจ้งซ่อม {ticket.ticket_number} เรียบร้อยแล้ว!', 'success')
        return redirect(url_for('tickets.ticket_detail', ticket_id=ticket.id))

    categories = Category.query.all()
    equipments = Equipment.query.all()
    return render_template(
        'ticket_create.html',
        categories=categories,
        equipments=equipments,
        branch_options=branch_options
    )


@tickets_bp.route('/<int:ticket_id>')
@login_required
def ticket_detail(ticket_id):
    """หน้ารายละเอียด Ticket"""
    ticket = Ticket.query.get_or_404(ticket_id)
    logs = ticket.logs.all()
    comments = ticket.comments.all()
    attachments = ticket.attachments.all()
    users = User.query.filter(User.role.in_(['Admin', 'Manager'])).all()

    # สร้าง QR Code
    qr_data = f"HELPDESK|{ticket.ticket_number}|{ticket.title}"
    qr_base64 = generate_qr_base64(qr_data)

    # คำนวณสถานะที่ผู้ใช้ปัจจุบันสามารถเปลี่ยนได้
    allowed_statuses = get_allowed_statuses(ticket.status, current_user.role)

    # ดึงข้อมูลค่าใช้จ่ายและอะไหล่
    expenses = ticket.expenses.all()
    total_expense = sum(e.total for e in expenses)
    spare_parts = SparePart.query.filter_by(is_active=True).order_by(SparePart.name).all()

    return render_template('ticket_detail.html',
                           ticket=ticket,
                           logs=logs,
                           comments=comments,
                           attachments=attachments,
                           assignable_users=users,
                           qr_base64=qr_base64,
                           allowed_statuses=allowed_statuses,
                           expenses=expenses,
                           total_expense=total_expense,
                           spare_parts=spare_parts)


@tickets_bp.route('/<int:ticket_id>/edit', methods=['GET', 'POST'])
@login_required
def ticket_edit(ticket_id):
    """แก้ไข Ticket"""
    ticket = Ticket.query.get_or_404(ticket_id)

    # ตรวจสอบสิทธิ์: เจ้าของ ticket หรือ Admin เท่านั้น
    if ticket.requester_id != current_user.id and current_user.role != 'Admin':
        flash('คุณไม่มีสิทธิ์แก้ไข Ticket นี้', 'danger')
        return redirect(url_for('tickets.ticket_detail', ticket_id=ticket.id))

    # ห้ามแก้ไข ticket ที่ปิดแล้ว
    if ticket.status in ['Closed', 'Resolved']:
        flash('ไม่สามารถแก้ไข Ticket ที่เสร็จสิ้นหรือปิดแล้ว', 'warning')
        return redirect(url_for('tickets.ticket_detail', ticket_id=ticket.id))

    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        description = request.form.get('description', '').strip()
        priority = request.form.get('priority', 'Medium')
        category_id = request.form.get('category_id', type=int)

        if not title:
            flash('กรุณากรอกหัวข้อปัญหา', 'danger')
            return redirect(url_for('tickets.ticket_edit', ticket_id=ticket.id))

        changes = []
        if ticket.title != title:
            changes.append(f'หัวข้อ: "{ticket.title}" → "{title}"')
        if ticket.priority != priority:
            changes.append(f'ความสำคัญ: {ticket.get_priority_thai()} → {Ticket(priority=priority).get_priority_thai()}')

        ticket.title = title
        ticket.description = description
        ticket.priority = priority
        ticket.category_id = category_id

        if changes:
            log = TicketLog(
                ticket_id=ticket.id,
                actor_id=current_user.id,
                action='แก้ไขข้อมูล',
                comment=', '.join(changes),
            )
            db.session.add(log)

        # เพิ่มไฟล์แนบใหม่
        files = request.files.getlist('attachments')
        for f in files:
            if f and f.filename and allowed_file(f.filename):
                ext = f.filename.rsplit('.', 1)[1].lower()
                safe_name = f"{uuid.uuid4().hex}.{ext}"
                save_path = os.path.join(current_app.config['UPLOAD_FOLDER'], safe_name)
                f.save(save_path)
                attachment = FileAttachment(
                    ticket_id=ticket.id, filename=safe_name,
                    original_name=f.filename, file_size=os.path.getsize(save_path),
                    mime_type=f.content_type, uploaded_by=current_user.id,
                )
                db.session.add(attachment)

        db.session.commit()
        flash('แก้ไข Ticket เรียบร้อย', 'success')
        return redirect(url_for('tickets.ticket_detail', ticket_id=ticket.id))

    categories = Category.query.all()
    return render_template('ticket_edit.html', ticket=ticket, categories=categories)


@tickets_bp.route('/<int:ticket_id>/delete', methods=['POST'])
@login_required
def ticket_delete(ticket_id):
    """ลบ Ticket (Admin เท่านั้น)"""
    if current_user.role != 'Admin':
        flash('เฉพาะ Admin เท่านั้นที่สามารถลบ Ticket ได้', 'danger')
        return redirect(url_for('tickets.ticket_list'))

    ticket = Ticket.query.get_or_404(ticket_id)
    ticket_number = ticket.ticket_number

    # ลบไฟล์แนบ
    for att in ticket.attachments.all():
        filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], att.filename)
        if os.path.exists(filepath):
            os.remove(filepath)

    # ลบข้อมูลที่เกี่ยวข้อง
    TicketComment.query.filter_by(ticket_id=ticket.id).delete()
    TicketLog.query.filter_by(ticket_id=ticket.id).delete()
    FileAttachment.query.filter_by(ticket_id=ticket.id).delete()
    db.session.delete(ticket)
    db.session.commit()

    flash(f'ลบ Ticket {ticket_number} เรียบร้อย', 'success')
    return redirect(url_for('tickets.ticket_list'))


@tickets_bp.route('/<int:ticket_id>/upload', methods=['POST'])
@login_required
def upload_file(ticket_id):
    """อัปโหลดไฟล์แนบเพิ่ม"""
    ticket = Ticket.query.get_or_404(ticket_id)

    files = request.files.getlist('attachments')
    count = 0
    for f in files:
        if f and f.filename and allowed_file(f.filename):
            ext = f.filename.rsplit('.', 1)[1].lower()
            safe_name = f"{uuid.uuid4().hex}.{ext}"
            save_path = os.path.join(current_app.config['UPLOAD_FOLDER'], safe_name)
            f.save(save_path)
            attachment = FileAttachment(
                ticket_id=ticket.id, filename=safe_name,
                original_name=f.filename, file_size=os.path.getsize(save_path),
                mime_type=f.content_type, uploaded_by=current_user.id,
            )
            db.session.add(attachment)
            count += 1

    if count > 0:
        db.session.commit()
        flash(f'อัปโหลดไฟล์ {count} ไฟล์เรียบร้อย', 'success')
    else:
        flash('ไม่พบไฟล์ที่ถูกต้อง กรุณาตรวจสอบนามสกุลไฟล์', 'warning')

    return redirect(url_for('tickets.ticket_detail', ticket_id=ticket.id))


@tickets_bp.route('/attachment/<int:attachment_id>/delete', methods=['POST'])
@login_required
def delete_attachment(attachment_id):
    """ลบไฟล์แนบ"""
    att = FileAttachment.query.get_or_404(attachment_id)
    ticket_id = att.ticket_id

    # ตรวจสอบสิทธิ์
    if att.uploaded_by != current_user.id and current_user.role != 'Admin':
        flash('คุณไม่มีสิทธิ์ลบไฟล์นี้', 'danger')
        return redirect(url_for('tickets.ticket_detail', ticket_id=ticket_id))

    filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], att.filename)
    if os.path.exists(filepath):
        os.remove(filepath)

    db.session.delete(att)
    db.session.commit()
    flash('ลบไฟล์แนบเรียบร้อย', 'success')
    return redirect(url_for('tickets.ticket_detail', ticket_id=ticket_id))


@tickets_bp.route('/<int:ticket_id>/comment', methods=['POST'])
@login_required
def add_comment(ticket_id):
    """เพิ่มความคิดเห็น"""
    ticket = Ticket.query.get_or_404(ticket_id)
    message = request.form.get('message', '').strip()

    if message:
        comment = TicketComment(
            ticket_id=ticket.id,
            user_id=current_user.id,
            message=message,
        )
        db.session.add(comment)
        db.session.commit()
        flash('เพิ่มความคิดเห็นเรียบร้อย', 'success')

    return redirect(url_for('tickets.ticket_detail', ticket_id=ticket.id))


@tickets_bp.route('/<int:ticket_id>/status', methods=['POST'])
@login_required
def change_status(ticket_id):
    """เปลี่ยนสถานะ Ticket (ตรวจสอบสิทธิ์ตาม Workflow)"""
    ticket = Ticket.query.get_or_404(ticket_id)
    new_status = request.form.get('status')
    comment = request.form.get('comment', '').strip()

    # ตรวจสอบสิทธิ์ตาม Workflow
    allowed = get_allowed_statuses(ticket.status, current_user.role)
    if new_status not in allowed:
        flash('คุณไม่มีสิทธิ์เปลี่ยนสถานะนี้', 'danger')
        return redirect(url_for('tickets.ticket_detail', ticket_id=ticket.id))

    old_status = ticket.status
    ticket.status = new_status

    if new_status == 'Resolved':
        ticket.resolved_at = datetime.utcnow()

    log = TicketLog(
        ticket_id=ticket.id,
        actor_id=current_user.id,
        action='เปลี่ยนสถานะ',
        old_value=old_status,
        new_value=new_status,
        comment=comment,
    )
    db.session.add(log)
    db.session.commit()

    # ส่งการแจ้งเตือน
    try:
        config = SystemConfig.get_notification_config()
        from ..utils import STATUS_LABELS
        action_text = f'เปลี่ยนสถานะเป็น "{STATUS_LABELS.get(new_status, new_status)}" โดย {current_user.name}'
        send_notification(ticket, action_text, config)
    except Exception:
        pass

    flash(f'เปลี่ยนสถานะเป็น "{ticket.get_status_thai()}" เรียบร้อย', 'success')
    return redirect(url_for('tickets.ticket_detail', ticket_id=ticket.id))


@tickets_bp.route('/<int:ticket_id>/assign', methods=['POST'])
@login_required
def assign_ticket(ticket_id):
    """มอบหมายงานให้ผู้รับผิดชอบ"""
    ticket = Ticket.query.get_or_404(ticket_id)
    assignee_id = request.form.get('assignee_id', type=int)

    if assignee_id:
        assignee = User.query.get(assignee_id)
        if assignee:
            ticket.assignee_id = assignee_id
            log = TicketLog(
                ticket_id=ticket.id,
                actor_id=current_user.id,
                action='มอบหมายงาน',
                new_value=assignee.name,
            )
            db.session.add(log)
            db.session.commit()
            flash(f'มอบหมายงานให้ {assignee.name} เรียบร้อย', 'success')

    return redirect(url_for('tickets.ticket_detail', ticket_id=ticket.id))


@tickets_bp.route('/import', methods=['GET', 'POST'])
@login_required
def import_csv():
    """นำเข้า Ticket จากไฟล์ CSV"""
    if current_user.role not in ['Admin', 'Manager']:
        flash('คุณไม่มีสิทธิ์นำเข้าข้อมูล', 'danger')
        return redirect(url_for('tickets.ticket_list'))

    if request.method == 'POST':
        file = request.files.get('csv_file')
        if not file or not file.filename.endswith('.csv'):
            flash('กรุณาอัปโหลดไฟล์ CSV', 'danger')
            return redirect(url_for('tickets.import_csv'))

        try:
            content = file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(content))
            count = 0

            category_map = {c.name: c.id for c in Category.query.all()}
            priority_map = {'ต่ำ': 'Low', 'ปานกลาง': 'Medium', 'สูง': 'High', 'วิกฤต': 'Critical',
                            'Low': 'Low', 'Medium': 'Medium', 'High': 'High', 'Critical': 'Critical'}

            for row in reader:
                title = row.get('หัวข้อ', row.get('title', '')).strip()
                if not title:
                    continue

                description = row.get('รายละเอียด', row.get('description', '')).strip()
                priority_raw = row.get('ความสำคัญ', row.get('priority', 'Medium')).strip()
                priority = priority_map.get(priority_raw, 'Medium')
                cat_name = row.get('หมวดหมู่', row.get('category', '')).strip()
                category_id = category_map.get(cat_name)

                ticket = Ticket(
                    ticket_number=Ticket.generate_ticket_number(),
                    title=title,
                    description=description,
                    priority=priority,
                    category_id=category_id,
                    requester_id=current_user.id,
                    status='Pending',
                )
                db.session.add(ticket)
                db.session.flush()

                log = TicketLog(
                    ticket_id=ticket.id, actor_id=current_user.id,
                    action='นำเข้าจาก CSV', new_value='Pending',
                )
                db.session.add(log)
                count += 1

            db.session.commit()
            flash(f'นำเข้า {count} รายการเรียบร้อย', 'success')
            return redirect(url_for('tickets.ticket_list'))

        except Exception as e:
            db.session.rollback()
            flash(f'เกิดข้อผิดพลาด: {str(e)}', 'danger')
            return redirect(url_for('tickets.import_csv'))

    return render_template('import_csv.html')


@tickets_bp.route('/export/csv')
@login_required
def export_csv():
    """Export รายการ Ticket เป็น CSV"""
    tickets = Ticket.query.order_by(Ticket.created_at.desc()).all()

    output = io.StringIO()
    output.write('\ufeff')  # BOM for Excel to read Thai
    writer = csv.writer(output)
    writer.writerow(['รหัส', 'หัวข้อ', 'หมวดหมู่', 'ความสำคัญ', 'สถานะ', 'ผู้แจ้ง', 'ผู้รับผิดชอบ', 'วันที่สร้าง'])

    for t in tickets:
        from ..utils import to_buddhist_era
        writer.writerow([
            t.ticket_number,
            t.title,
            t.category.name if t.category else '-',
            t.get_priority_thai(),
            t.get_status_thai(),
            t.requester.name if t.requester else '-',
            t.assignee.name if t.assignee else '-',
            to_buddhist_era(t.created_at),
        ])

    response = Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=tickets_export.csv'}
    )
    return response


@tickets_bp.route('/export/excel')
@login_required
def export_excel():
    """Export รายการ Ticket เป็น Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    tickets = Ticket.query.order_by(Ticket.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"

    # Header
    headers = ['รหัส', 'หัวข้อ', 'หมวดหมู่', 'ความสำคัญ', 'สถานะ', 'ผู้แจ้ง', 'ผู้รับผิดชอบ', 'วันที่สร้าง']
    header_fill = PatternFill(start_color='B91C1C', end_color='B91C1C', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Data
    from ..utils import to_buddhist_era
    for row, t in enumerate(tickets, 2):
        ws.cell(row=row, column=1, value=t.ticket_number)
        ws.cell(row=row, column=2, value=t.title)
        ws.cell(row=row, column=3, value=t.category.name if t.category else '-')
        ws.cell(row=row, column=4, value=t.get_priority_thai())
        ws.cell(row=row, column=5, value=t.get_status_thai())
        ws.cell(row=row, column=6, value=t.requester.name if t.requester else '-')
        ws.cell(row=row, column=7, value=t.assignee.name if t.assignee else '-')
        ws.cell(row=row, column=8, value=to_buddhist_era(t.created_at))

    # Auto-width
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_length + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=tickets_export.xlsx'}
    )
